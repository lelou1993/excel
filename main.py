from openpyxl.styles import PatternFill
from openpyxl.styles import colors
from openpyxl.styles import Font
import openpyxl as vb

workbook_a = vb.load_workbook(r'C:\Users\asus\Desktop\系统清单2022_V1 (7).xlsx')
workbook_b = vb.load_workbook(r'C:\Users\asus\Desktop\系统清单2022_V1 (5).xlsx')
sheet_a = workbook_a['业务模块']
sheet_b = workbook_b['业务模块']
maxrow = sheet_a.max_row
maxcolumn = sheet_b.max_column
for i in range(1, maxrow):
    for j in range(1, maxcolumn):
        cell_a = sheet_a.cell(i, j)
        cell_b = sheet_b.cell(i, j)

        if cell_a.value != cell_b.value:
            cell_a.fill = PatternFill("solid", fgColor="FFFF00")
            cell_a.font = Font(color=colors.BLUE, bold=True)
            cell_b.fill = PatternFill("solid", fgColor="FFFF00")
            cell_b.font = Font(color=colors.BLUE, bold=True)
workbook_a.save("1.xlsx")
workbook_b.save("2.xlsx")